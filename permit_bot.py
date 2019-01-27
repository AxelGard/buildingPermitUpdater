#!/usr/bin/python3.6
import os, sys, re, time, random, datetime

#Excel libs (openpyxl)
from openpyxl import load_workbook as oplw
from openpyxl.styles import PatternFill, Color, colors
import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import Color
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl import workbook

#email lib
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

#clear screen
#os.system('clear')
body = ""
ExpiredToF = True

def setUpEmail():
    #add cell values in email
    bildPos = 'C' + str(caounterForColor)
    bildning = sh[bildPos].value
    bildDatePos = 'G' + str(caounterForColor)
    EndDate = sh[bildDatePos].value
    EndDate = EndDate.strftime("%Y-%m-%d")
    global body
    global ExpiredToF

    #if going to Expire -> == Flase if it have -> == True
    if ExpiredToF == True:
        body = body + "\n " + str(bildning) + " building permit has expired : " + str(EndDate) + "\n"
    else:
        body = body + "\n " + str(bildning) + " building permit will soon expire at : " + str(EndDate) + "\n"

def SendEmail():
    #basic info
    senderEmail = 'excel.bot@gmail.com' #bot gmail
    senderPassword = 'BotPassword' #bot password
    receverEmail = 'youre@email.com' #result will be send to this emails
    subjectEmail = 'building permit expiration dates' #subject on email

    global body

    #Subjekts, from and to
    msg = MIMEMultipart()
    msg['From'] = senderEmail
    msg['To'] = receverEmail
    msg['Subject'] = subjectEmail

    #take text to email format
    msg.attach(MIMEText(body,'plain'))
    text = msg.as_string()

    #server conection
    server = smtplib.SMTP('smtp.gmail.com',587)
    server.starttls()
    server.login(senderEmail, senderPassword)

    #sending email
    server.sendmail(senderEmail, receverEmail, text)
    server.quit()


#get today's date in multi dem array
TodaysDate = []
#year
TodaysDateYear = time.strftime("%Y")
TodaysDateYear = int(TodaysDateYear)
#Month
TodaysDateMonth = time.strftime("%m")
TodaysDateMonth = int(TodaysDateMonth)
#Day
TodaysDateDay = time.strftime("%d")
TodaysDateDay = int(TodaysDateDay)
#add all them to 2d list
TodaysDate = [TodaysDateYear, TodaysDateMonth, TodaysDateDay]

#get values from excel file
#openpyxl
wb = oplw("Excel.xlsx", data_only=True)
#sh = wb["Blad1"]
sh = wb.active

XlValueList = []
cl = sh['G'] #cl were all dates are stord in excel doc

Excel = []

#runs thrue all values in cl and appedns to a list
for x in range(len(cl)):
    XlValueList.append(cl[x].value)
XlValueList.remove('Tom')
#puts all values in a 2d list
for y in XlValueList:
    tempList = []
    #Year
    tempYear = y.strftime('%Y')
    tempYear = int(tempYear)
    tempList.append(tempYear)
    #Month
    tempMonth = y.strftime('%m')
    tempMonth = int(tempMonth)
    tempList.append(tempMonth)
    #day
    tempDay = y.strftime('%d')
    tempDay = int(tempDay)
    tempList.append(tempDay)
    #puts list in list -> 2d list
    Excel.append(tempList)


#check date vs etch value in list
caounterForColor = 1
for i in Excel:
    for j in i:
        caounterForColor = caounterForColor + 1
        # Check years
        if int(j) < int(TodaysDate[0]):
            #print("Expired : ", i, " pga : ", j)
            colorPos = 'G' + str(caounterForColor)
            c = sh[colorPos]
            c.font = Font(size=9, color="FF0000")#Set color
            break
        # if year is equal
        elif int(i[0]) == int(TodaysDate[0]):
            # Check month
            if int(i[1]) < int(TodaysDate[1]):
                #print("Expired : ", i, " pga : ", i[1])
                colorPos = 'G' + str(caounterForColor)
                c = sh[colorPos]
                c.font = Font(size=9, color="FF0000")#Set color
                break
            elif int(i[1]) > int(TodaysDate[1]) and int(i[1]) <= int(TodaysDate[1]) + 5 and int(j) == int(TodaysDate[0]):
                #print("Expires soon : ", i, " pga : ", i[1])
                colorPos = 'G' + str(caounterForColor)
                c = sh[colorPos]
                c.font = Font(size=9, color="FFBB00")#Set color
                ExpiredToF = False #will expire
                setUpEmail()
                break
            # if month is equal
            elif int(i[1]) == int(TodaysDate[1]):
                # Check day
                if int(i[2]) < int(TodaysDate[2]):
                    #print("Expired : ", i, " pga : ", i[2])
                    colorPos = 'G' + str(caounterForColor)
                    c = sh[colorPos]
                    c.font = Font(size=9, color="FF0000")#Set color
                    setUpEmail() #send email lower then day
                    break
                elif int(i[2]) > int(TodaysDate[2]):
                    #print("Expires soon : ", i, " pga : ", i[1])
                    colorPos = 'G' + str(caounterForColor)
                    c = sh[colorPos]
                    c.font = Font(size=9, color="FFBB00")#Set color
                    ExpiredToF = False # will expire
                    setUpEmail()
                    break
                else:
                    break
            else:
                break
        else:
            break
    ExpiredToF = True

wb.save('Excel.xlsx') #saves changes made to doc
if body != "":
    SendEmail()
#print("\nProgram is done ")
